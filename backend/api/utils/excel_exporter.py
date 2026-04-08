import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def export_timetables_to_excel(timetables: list, format_type: str = "tabs"):
    """
    timetables: list of dicts with:
        - entity_name: str
        - context: str ('group', 'teacher', 'room')
        - grid_data: dict
        - periods: list
        - assignments: list of tuples (subj_abbr, subj_name, teach_abbr, teach_name)
    format_type: 'tabs' or 'vertical'
    """
    wb = Workbook()
    
    # Remove the default sheet first unless we are doing vertical
    default_sheet = wb.active
    if format_type == "tabs":
        wb.remove(default_sheet)
    else:
        default_sheet.title = "All Timetables"
        
    current_row = 1
    
    for tt in timetables:
        if format_type == "tabs":
            ws = wb.create_sheet(title=f"{tt['entity_name']}"[:31]) # excel tab limit is 31 chars
            start_row = 1
        else:
            ws = wb.active
            start_row = current_row
            
        next_row = _write_timetable_to_sheet(
            ws=ws,
            start_row=start_row,
            grid_data=tt["grid_data"],
            periods=tt["periods"],
            context=tt["context"],
            entity_name=tt["entity_name"],
            assignments=tt.get("assignments", [])
        )
        
        if format_type == "vertical":
            current_row = next_row + 4 # skip some rows before the next table
            
    # If timetables was empty and tabs, we need at least one sheet
    if not wb.sheetnames:
        wb.create_sheet(title="No Data")
        
    # Return workbook bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _write_timetable_to_sheet(ws, start_row, grid_data, periods, context, entity_name, assignments):
    # Styling constants
    title_font = Font(bold=True, size=16, color="333333")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 0. Write Entity Title
    ws.cell(row=start_row, column=1, value=f"Timetable: {entity_name}").font = title_font
    current_row = start_row + 2
    
    # 1. Build Header Row (Days down the side, Periods across the top)
    ws.cell(row=current_row, column=1, value="Day / Period").font = header_font
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=1).alignment = center_align
    ws.cell(row=current_row, column=1).border = thin_border
    ws.column_dimensions['A'].width = 15

    for col_idx, p in enumerate(periods, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=f"P{p['period']}\n({p['start']}-{p['end']})")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        # Set column width via letter
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = 25
        
    current_row += 1

    # 2. Build Grid Data
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    days_present = [d for d in days if d in grid_data]

    for day in days_present:
        ws.cell(row=current_row, column=1, value=day).font = Font(bold=True)
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        
        row_cells = grid_data[day]
        for col_idx, cell_info in enumerate(row_cells, start=2):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align
            
            if cell_info is None:
                continue

            if cell_info.get("isContinuation"):
                cell.value = "-->"
                continue

            # Construct cell content based on context
            parts = []
            if cell_info.get("subject"):
                parts.append(cell_info["subject"])
            if context != "group" and cell_info.get("group"):
                parts.append(f"Grp: {cell_info['group']}")
            if context != "teacher" and cell_info.get("teacher"):
                parts.append(f"Prof: {cell_info['teacher']}")
            if context != "room" and cell_info.get("room"):
                parts.append(f"Rm: {cell_info['room']}")

            cell.value = "\n".join(parts)
            cell.fill = cell_fill
            
        current_row += 1

    # 3. Build Assignments Legend below the grid
    current_row += 2
    if assignments:
        ws.cell(row=current_row, column=1, value="Course Assignments & Legends").font = Font(bold=True, size=14)
        current_row += 1
        
        headers = ["Subject Abbr", "Subject Name", "Teacher Abbr", "Teacher Name"]
        for col_idx, text in enumerate(headers, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=text)
            c.font = Font(bold=True)
            c.border = thin_border
            c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            c.alignment = left_align
            
        current_row += 1
        
        # sort assignments by subject
        sorted_assignments = sorted(list(set(assignments)), key=lambda x: str(x[0]))
        
        for subj_abbr, subj_name, teach_abbr, teach_name in sorted_assignments:
            row_data = [subj_abbr, subj_name, teach_abbr, teach_name]
            for col_idx, val in enumerate(row_data, start=1):
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.border = thin_border
                c.alignment = left_align
            current_row += 1

    return current_row
