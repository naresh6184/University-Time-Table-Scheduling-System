import io
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from backend.database.models.student import StudentModel
from backend.database.models.branch import BranchModel
from backend.api.schemas.student_schema import StudentCreate

def preview_student_import(db: Session, file_content: bytes):
    """
    Parses the excel file and returns a summary of valid/invalid students.
    """
    try:
        wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
        ws = wb.active
    except Exception as e:
        return {"error": f"Failed to read Excel file: {str(e)}"}

    # Expected Columns: Name, Roll No, Program, Batch, Branch, Email
    # We'll allow some flexibility in headers
    headers = [cell.value for cell in ws[1]]
    
    # Map headers to internal keys
    header_map = {}
    for i, h in enumerate(headers):
        if not h: continue
        h_low = h.lower().strip()
        # Check 'branch' before 'id'/'name' to avoid 'branch_name' being misclassified as 'name'
        if 'branch' in h_low: header_map['branch_name'] = i
        elif 'name' in h_low: header_map['name'] = i
        elif 'roll' in h_low or ('id' in h_low and 'branch' not in h_low): header_map['student_id'] = i
        elif 'program' in h_low: header_map['program'] = i
        elif 'batch' in h_low: header_map['batch'] = i
        elif 'email' in h_low: header_map['email'] = i

    required = ['name', 'student_id', 'branch_name', 'program', 'batch']
    missing = [r for r in required if r not in header_map]
    if missing:
        return {"error": f"Missing required columns: {', '.join(missing)}"}

    # Pre-fetch branches for mapping
    branches = db.query(BranchModel).all()
    branch_map = {b.name.lower().strip(): b.branch_id for b in branches}
    branch_abbr_map = {b.abbreviation.lower().strip(): b.branch_id for b in branches if b.abbreviation}

    students_to_import = []
    invalid_rows = []
    
    # Check existing roll numbers to avoid duplicates
    existing_ids = {s[0] for s in db.query(StudentModel.student_id).all()}
    seen_in_file = set()  # Track duplicates within the file itself

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue # Skip empty rows

        data = {}
        for key, col_idx in header_map.items():
            data[key] = row[col_idx]

        errors = []
        
        # Validation
        name = str(data.get('name') or '').strip()
        student_id = str(data.get('student_id') or '').strip()
        branch_name = str(data.get('branch_name') or '').strip().lower()
        program = str(data.get('program') or '').strip()
        batch_val = data.get('batch')
        email = str(data.get('email') or '').strip()

        if not name: errors.append("Name is required")
        if not student_id: 
            errors.append("Roll Number/ID is required")
        elif student_id in existing_ids:
            errors.append(f"Roll Number {student_id} already exists in database")
        elif student_id in seen_in_file:
            errors.append(f"Duplicate Roll Number {student_id} in file")
        else:
            seen_in_file.add(student_id)

        if not program:
            errors.append("Program is required (e.g. B.Tech)")
        
        if not batch_val:
            errors.append("Batch is required (e.g. 2022)")

        # Smart Branch mapping
        # First, check if the value is a numeric branch_id
        branch_raw = data.get('branch_name')
        branch_id = None
        try:
            numeric_id = int(branch_raw)
            # Verify this branch_id exists
            if numeric_id in {b.branch_id for b in branches}:
                branch_id = numeric_id
        except (TypeError, ValueError):
            pass
        
        if not branch_id:
            branch_id = branch_map.get(branch_name) or branch_abbr_map.get(branch_name)
        
        # Substring matching (handles formats like "Name (Abbr)")
        if not branch_id:
            for abbr, b_id in branch_abbr_map.items():
                if abbr and f"({abbr})" in branch_name:
                    branch_id = b_id
                    break
        
        if not branch_id:
            for abbr, b_id in branch_abbr_map.items():
                if abbr and abbr == branch_name:
                    branch_id = b_id
                    break

        if not branch_id:
            for name_key, b_id in branch_map.items():
                if name_key and name_key in branch_name:
                    branch_id = b_id
                    break

        if not branch_id:
            if branch_name:
                errors.append(f"Branch '{data.get('branch_name')}' not found")
            else:
                errors.append("Branch is required")

        # Batch validation
        batch = None
        if batch_val:
            try:
                batch = int(batch_val)
            except:
                errors.append("Batch must be a number")

        if errors:
            invalid_rows.append({
                "line": row_idx,
                "name": name,
                "student_id": student_id,
                "errors": errors
            })
        else:
            students_to_import.append({
                "student_id": student_id,
                "name": name,
                "branch_id": branch_id,
                "batch": batch,
                "email": email,
                "program": program
            })

    return {
        "total": len(students_to_import) + len(invalid_rows),
        "valid_count": len(students_to_import),
        "invalid_count": len(invalid_rows),
        "valid_students": students_to_import,
        "invalid_rows": invalid_rows
    }

def commit_student_import(db: Session, students_data: list):
    """
    Inserts validated students into the database.
    """
    count = 0
    for s_data in students_data:
        # Check one last time for duplicate (to be safe against concurrent requests)
        exists = db.query(StudentModel).filter_by(student_id=s_data['student_id']).first()
        if not exists:
            student = StudentModel(**s_data)
            db.add(student)
            count += 1
    
    db.commit()
    return count
